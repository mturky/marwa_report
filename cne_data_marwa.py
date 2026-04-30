# %% [markdown]
# # CNE Sales Report Generator
# This notebook processes **Financial Transactions (FT)** and **New Sales** data,
# merges them with **BeinData** reference information, and exports separate
# Excel reports for **Dealers** and **Direct Sales**.

# %%
# ── Imports ─────────────────────────────────────────────────────────────────
import pandas as pd
from datetime import datetime, timedelta
import calendar
import shutil
import warnings

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from functions import create_folder_if_not_exists, search_files,sendMail

# Suppress noisy warnings (e.g. pandas FutureWarnings)
warnings.filterwarnings("ignore")

# %%
# ── Configuration / Parameters ──────────────────────────────────────────────

# Date helpers
date = datetime.now().date()
base_dir_format = datetime.now().strftime('%d.%m.%y')
base_dir = f's:\\\\{base_dir_format}'
    
# Output directory (created if missing)
result_directory = 's:\\\\playground'
create_folder_if_not_exists(result_directory)

# Pandas display settings
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', 50)
pd.set_option('display.float_format', '{:.2f}'.format)

if date.day == 1:  # First day of month -> report on entire previous month
    last_month = date.month - 1 if date.month > 1 else 12
    last_month_year = date.year if date.month > 1 else date.year - 1
    FROM_DATE = pd.Timestamp(last_month_year, last_month, 1)
    TO_DATE = pd.Timestamp(last_month_year, last_month, calendar.monthrange(last_month_year, last_month)[1])

elif date.weekday() == 6:  # Sunday
    FROM_DATE = pd.Timestamp(date - timedelta(days=3))
    TO_DATE = pd.Timestamp(date - timedelta(days=1))

else:
    FROM_DATE = pd.Timestamp(date - timedelta(days=1))
    TO_DATE = pd.Timestamp(date - timedelta(days=1))


# FROM_DATE = "2026-02-01"
# TO_DATE   = "2026-02-26"

# Output file names
OUTPUT_FILE_DEALERS = f'Dealers report {FROM_DATE:%Y-%m-%d} to {TO_DATE:%Y-%m-%d}.xlsx'
OUTPUT_FILE_DS      = f'Direct Sales report {FROM_DATE:%Y-%m-%d} to {TO_DATE:%Y-%m-%d}.xlsx'

# %%
# ── Locate Source Files ─────────────────────────────────────────────────────

# Partial file-name patterns used to find the CSVs in today's directory
FT_PARTIAL_NAME        = 'NEWCNEFINTRANSRPT'
NEWSALES_PARTIAL_NAME  = 'CNENEWCAPTURERPT.CSV'
BEINDATA_PARTIAL_NAME  = 'BEINDATANEWRPT'

ft_found_files        = search_files(base_dir, FT_PARTIAL_NAME)
newsales_found_files  = search_files(base_dir, NEWSALES_PARTIAL_NAME)
beindata_found_files  = search_files(base_dir, BEINDATA_PARTIAL_NAME)

# ── Validate that all required files were found ────────────────────────────
missing_files = []
if not ft_found_files:
    missing_files.append(f'FT file ({FT_PARTIAL_NAME})')
if not newsales_found_files:
    missing_files.append(f'New-sales file ({NEWSALES_PARTIAL_NAME})')
if len(beindata_found_files) < 2:
    missing_files.append(f'BeinData file ({BEINDATA_PARTIAL_NAME}) — found {len(beindata_found_files)}, need at least 2')

if missing_files:
    error_msg = f'Missing source files in {base_dir}:\n' + '\n'.join(f'  - {f}' for f in missing_files)
    print(f'ERROR: {error_msg}')
    sendMail(
        subject='CNE Report - Missing Source Files',
        cc=['nwagih@cne.com.eg','dsabry@cne.com.eg'],
        body=f'<p>The CNE Sales Report could not be generated.<br><br><b>Missing files:</b></p><ul>' +
             ''.join(f'<li>{f}</li>' for f in missing_files) +
             f'</ul><p>Search directory: <code>{base_dir}</code></p>',
    )
    raise SystemExit(error_msg)

print(f'FT file:        {ft_found_files[0]}')
print(f'New-sales file: {newsales_found_files[0]}')
print(f'BeinData file:  {beindata_found_files[1]}')
print(f'Current date:   {date}')

# %% [markdown]
# ## Helper Functions

# %%
def clean_text(text: str) -> str:
    """Remove trailing name tokens from a user-fullname string to derive a location.

    If the last word is a single character (initial) and there are ≥ 2 words,
    drop the last *two* words; otherwise drop only the last word.
    """
    words = text.split()
    if len(words) >= 2 and len(words[-1]) == 1:
        return ' '.join(words[:-2])
    return ' '.join(words[:-1])


def diff_month(d1: pd.Series, d2: pd.Series) -> pd.Series:
    """Return the difference in whole months between two datetime Series."""
    return (d1.dt.year - d2.dt.year) * 12 + d1.dt.month - d2.dt.month


def format_excel_table(file_path: str, table_name: str) -> None:
    """Add an Excel Table with styling and auto-fit column widths.

    Reads an existing workbook, converts the 'Data' sheet into a named
    table with striped rows, then auto-fits every column.
    """
    wb = load_workbook(file_path)
    ws = wb['Data']

    end_row = ws.max_row
    end_col = ws.max_column

    # Create and style the table
    table = Table(
        displayName=table_name,
        ref=f'A1:{get_column_letter(end_col)}{end_row}',
    )
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Auto-fit column widths (with a small padding)
    for col_idx in range(1, end_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            (len(str(cell.value)) for cell in ws[col_letter] if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = max_length + 5

    wb.save(file_path)

# %% [markdown]
# ## Load & Parse Raw Data

# %%
# ── Load CSV files (all columns as strings to avoid type-coercion issues) ──
ft_raw       = pd.read_csv(ft_found_files[0],       dtype='str', on_bad_lines='skip')
newsales_raw = pd.read_csv(newsales_found_files[0],  dtype='str')
beindata_raw = pd.read_csv(beindata_found_files[1],  dtype='str')

# ── Parse date columns ──────────────────────────────────────────────────────
newsales_raw['Customer Created Date'] = pd.to_datetime(
    newsales_raw['Customer Created Date'], dayfirst=True
)
ft_raw['Created Date'] = pd.to_datetime(ft_raw['Created Date'], dayfirst=True)

# %%
ft_raw.loc[ft_raw['Subscriber Nr']=='15515110']

# %% [markdown]
# ## Prepare Financial Transactions (FT)

# %%


# %%
ft = ft_raw.copy()

# Keep only posted JV and Payment documents within the reporting period
ft = ft.loc[
    (ft['Doc Status'] == 'Posted')
    & (ft['Created Date'] >= FROM_DATE)
    & (ft['Created Date'] <= TO_DATE)
    & (ft['Doc Type'].isin(['JV', 'Payment']))
]

# Derive location from user full-name
ft['Location']      = ft['User Fullname'].apply(clean_text)
ft['Location_Temp'] = ft['Collecting Entity']


# Enrich with decoder info from BeinData (Smart Card → Decoder, SC, STB type)
beindata_ref = beindata_raw[['Decoder', 'Smart Card', 'Item Description STB']].copy()

ft = ft.merge(
    beindata_ref,
    left_on='Smartcard',
    right_on='Smart Card',
    how='left',
)

# %% [markdown]
# ## Prepare New Sales

# %%
newsales = newsales_raw.copy()

# Filter: reporting period, non-null box numbers/contracts, dealer prefix
newsales = newsales.loc[
    (newsales['Customer Created Date'] >= FROM_DATE)
    & (newsales['Customer Created Date'] <= TO_DATE)
    & newsales['Box Number'].notna()
    & newsales['Contract Status'].notna()
    & newsales['Dealer Code'].str.lower().str.startswith('bein')
]

# ── Fill missing payment amounts from FT ────────────────────────────────────
has_amount    = newsales['Payment Amount'].notna()
ns_with_amt   = newsales.loc[has_amount]
ns_no_amt     = newsales.loc[~has_amount]

# Left-join to FT on subscriber + date to pull the amount
ns_no_amt = ns_no_amt.merge(
    ft[['Subscriber Nr', 'Created Date', 'Amount']],
    left_on=['Customer Number', 'Customer Created Date'],
    right_on=['Subscriber Nr', 'Created Date'],
    how='left',
).drop(columns=['Subscriber Nr', 'Created Date', 'Payment Amount']
).rename(columns={'Amount': 'Payment Amount'})

# Recombine
newsales = pd.concat([ns_no_amt, ns_with_amt], ignore_index=True).drop_duplicates(keep='first')

# ── Classify source & type ──────────────────────────────────────────────────
newsales['Sale Source'] = 'Dealers'
newsales['Sale Type']   = 'New Sale'
newsales.loc[newsales['Dealer Type'] == 'Partner Head Office', 'Sale Source'] = 'Direct Sales'

newsales['dataset']       = 'ft_clone'
newsales['Location']      = newsales['Dealer Name']
newsales['Location_Temp'] = ''
newsales['Staff #']       = ''

# ── Enrich with BeinData (decoder type) ─────────────────────────────────────
newsales = newsales.merge(
    beindata_ref,
    left_on='Box Number',
    right_on='Decoder',
    how='left',
).drop(columns=['Decoder'])

# ── Enrich with FT (MOP, FTNR, Bill Period) ────────────────────────────────
newsales = newsales.merge(
    ft[['Subscriber Nr', 'Created Date', 'Amount', 'Pay Mode', 'Ftnr', 'Bill Period']],
    left_on=['Customer Number', 'Customer Created Date', 'Payment Amount'],
    right_on=['Subscriber Nr', 'Created Date', 'Amount'],
    how='left',
).drop(columns=['Subscriber Nr', 'Created Date', 'Amount'])

# Default pay mode to Cash
newsales['Pay Mode'] = newsales['Pay Mode'].fillna('Cash')

# Parse contract dates & set contract period
newsales['Start Date'] = pd.to_datetime(newsales['Start Date'], dayfirst=True)
newsales['End Date']   = pd.to_datetime(newsales['End Date'],   dayfirst=True)
newsales['Contract Period'] = '12M'

# %% [markdown]
# ## Split FT into Dealers & Direct Sales Renewals

# %%
# ── Dealers (entity type = 'Bein Dealer', doc type = 'JV') ──────────────────
ft_dealers = ft.loc[
    (ft['Default Entity Type'] == 'Bein Dealer') & (ft['Doc Type'] == 'JV')
].copy()
ft_dealers['Sale Source'] = 'Dealers'
ft_dealers['Sale Type']   = 'Renewal'
ft_dealers['dataset']     = 'ft_clone_dealers'

# ── Direct Sales (entity type = 'Partner Head Office', doc type = 'Payment') ─
ft_ds = ft.loc[
    (ft['Default Entity Type'] == 'Partner Head Office') & (ft['Doc Type'] == 'Payment')
].copy()
ft_ds['Sale Source'] = 'Direct Sales'
ft_ds['Sale Type']   = 'Renewal'
ft_ds['dataset']     = 'ft_clone_ds'

# %% [markdown]
# ## Standardise Columns & Merge All Data

# %%
# Unified column order for the final report
REPORT_COLS = [
    'TRANSACTION DATE', 'LOCATION', 'AGENT', 'SERIAL', 'PACKAGE','PAYMENT PLAN',
    'MOP', 'BOX TYPE', 'CSD', 'CED', 'AMOUNT', 'STAFF #', 'FTNR',
    'CONTRACT PERIOD', 'SALE SOURCE', 'SALE TYPE', 'Location_Temp',
]

# ── New Sales → standard columns ────────────────────────────────────────────
newsales_final = newsales[[
    'Customer Created Date', 'Location', 'User Name', 'Smart Card', 'Plan','Bill Frequency',
    'Pay Mode', 'Item Description STB', 'Start Date', 'End Date',
    'Payment Amount', 'Staff #', 'Ftnr', 'Contract Period',
    'Sale Source', 'Sale Type', 'Location_Temp',
]].copy()
newsales_final['Contract Period'] = '12M'
newsales_final.columns = REPORT_COLS
newsales_final.to_csv('newsales.csv', index=False)

# ── FT (Dealers + DS) → standard columns ───────────────────────────────────
ft_combined = pd.concat([ft_dealers, ft_ds], ignore_index=True)
ft_combined['Start Date']       = ''
ft_combined['End Date']         = ''
ft_combined['Staff #']          = ''
ft_combined['Contract Period']  = '12M'
ft_combined['Payment Plan']  = ''


ft_final = ft_combined[[
    'Created Date', 'Location', 'User Name', 'Smartcard', 'Plan Name','Payment Plan',
    'Pay Mode', 'Item Description STB', 'Start Date', 'End Date',
    'Amount', 'Staff #', 'Ftnr', 'Contract Period',
    'Sale Source', 'Sale Type', 'Location_Temp',
]].drop_duplicates()
ft_final['Pay Mode'] = ft_final['Pay Mode'].fillna('Cash')
ft_final.columns = REPORT_COLS
ft_final.to_csv('ft_final.csv', index=False)

# %%
ft_final

# %% [markdown]
# ## Final Assembly & Export

# %%
# ── Combine all data ────────────────────────────────────────────────────────
all_final = (
    pd.concat([ft_final, newsales_final], ignore_index=True)
    .drop_duplicates()
)
all_final['TRANSACTION DATE'] = pd.to_datetime(
    all_final['TRANSACTION DATE'], dayfirst=True
)

# For Direct-Sales renewals, override LOCATION with the collecting entity
mask_ds_renewal = (
    (all_final['SALE TYPE'] == 'Renewal') & (all_final['SALE SOURCE'] == 'Direct Sales')
)
all_final.loc[mask_ds_renewal, 'LOCATION'] = all_final.loc[mask_ds_renewal, 'Location_Temp']

# Drop helper column & add empty PACKAGE placeholder
all_final = all_final.drop(columns=['Location_Temp'])
# all_final['PACKAGE'] = ''

# Reorder & sort chronologically
FINAL_COLS = [
    'TRANSACTION DATE', 'LOCATION', 'AGENT', 'SERIAL', 'PACKAGE',
    'PAYMENT PLAN', 'MOP', 'SALE TYPE', 'BOX TYPE', 'CSD', 'CED',
    'AMOUNT', 'STAFF #', 'FTNR', 'CONTRACT PERIOD', 'SALE SOURCE',
]
all_final = all_final[FINAL_COLS].sort_values('TRANSACTION DATE')

# ── Split by sale source & drop that column from each subset ────────────────
all_final_dealers = all_final.loc[
    all_final['SALE SOURCE'] == 'Dealers'
].drop(columns='SALE SOURCE')

all_final_ds = all_final.loc[
    all_final['SALE SOURCE'] == 'Direct Sales'
].drop(columns='SALE SOURCE')

# ── Export to Excel with table formatting ───────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE_DEALERS, engine='openpyxl') as writer:
    all_final_dealers.to_excel(writer, sheet_name='Data', index=False)
format_excel_table(OUTPUT_FILE_DEALERS, 'DataTableDealers')

with pd.ExcelWriter(OUTPUT_FILE_DS, engine='openpyxl') as writer:
    all_final_ds.to_excel(writer, sheet_name='Data', index=False)
format_excel_table(OUTPUT_FILE_DS, 'DataTableDS')

print(f'✅ Dealers report saved to:      {OUTPUT_FILE_DEALERS}')
print(f'✅ Direct Sales report saved to: {OUTPUT_FILE_DS}')

# %%
htmlbody = """
<html>
<head>
    <style>
        td,th {
            padding: 5px;
        }
        p{
            font-size: 14px;
        }
        .signature{
            font-size: 16px;
        }
    </style>
</head>
<body>
    <p>Dear all</p>
    <p > Kindly find the attached reports</br>
       
    </p>


    <p class="signature">
    MTURKY
    </p
    </html>
    """
    
    
to_list = [
    'ozturkha@bein.com'
]

cc_list = [
    'sahinlerb@bein.com',
    'iyengarb@bein.com',
    'hassanm@bein.com',
    'tmohamed@cne.com.eg',
    'nwagih@cne.com.eg',
    'mturky@cne.com.eg'
]

sendMail('CNE Data',body=htmlbody,to=to_list,cc=cc_list,attachments=[OUTPUT_FILE_DEALERS,OUTPUT_FILE_DS])


